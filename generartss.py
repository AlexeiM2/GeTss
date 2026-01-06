# app.py
# Generador Excel por Edificio
# LISTADO GLOBAL + IMÁGENES + FECHA + ZIP FINAL

import streamlit as st
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.comments import Comment
from io import BytesIO
import zipfile
from copy import copy

st.set_page_config(page_title="Generador Excel por Edificio", layout="wide")

st.title("📊 Generador de Excel por Dependencia y Edificio")

# --- Carga del modelo ---
modelo_file_path = "Modelo.xlsx"
wb = load_workbook(modelo_file_path)

# --- Dependencia ---
st.header("2️⃣ Datos generales")
dependencia = st.text_input("Nombre de la dependencia")
fecha = st.text_input("Fecha (ej: 16/09/2026)")

# --- Edificios ---
st.header("3️⃣ Edificios")

if "edificios" not in st.session_state:
    st.session_state.edificios = []

col1, col2 = st.columns([3, 1])
with col1:
    edificio = st.text_input("Nombre del edificio")
with col2:
    if st.button("➕ Agregar"):
        if edificio:
            st.session_state.edificios.append({
                "nombre": edificio,
                "tipo": "Ambos",
                "imagenes_ap": [],
                "imagen_tss": None,
                "cantidad_switches": 1,
                "cantidad_aps": 1,
                "estados_switches": ["Reemplazo"],
                "estados_aps": ["Nuevo"]
            })

st.subheader("Listado de edificios")

for i, e in enumerate(st.session_state.edificios):
    with st.expander(f"🏢 {e['nombre']}", expanded=True):
        st.write("### Configuración General")
        c1, c2, c3 = st.columns([2, 2, 1])
        
        # Cantidad de equipos y tipo de hojas
        with c1:
            cantidad_sw = st.number_input(
                "Cantidad de Switches",
                min_value=0,
                value=e.get("cantidad_switches", 1),
                key=f"cant_sw_{i}"
            )
            
            cantidad_ap = st.number_input(
                "Cantidad de APs",
                min_value=0,
                value=e.get("cantidad_aps", 1),
                key=f"cant_ap_{i}"
            )
            
            # Actualizar cantidad y ajustar listas de estados
            if cantidad_sw != e.get("cantidad_switches", 1):
                st.session_state.edificios[i]["cantidad_switches"] = cantidad_sw
                # Ajustar lista de estados
                estados_actuales = e.get("estados_switches", [])
                if len(estados_actuales) < cantidad_sw:
                    # Agregar más estados
                    estados_actuales.extend(["Reemplazo"] * (cantidad_sw - len(estados_actuales)))
                else:
                    # Recortar estados
                    estados_actuales = estados_actuales[:cantidad_sw]
                st.session_state.edificios[i]["estados_switches"] = estados_actuales
            
            if cantidad_ap != e.get("cantidad_aps", 1):
                st.session_state.edificios[i]["cantidad_aps"] = cantidad_ap
                # Ajustar lista de estados
                estados_actuales = e.get("estados_aps", [])
                if len(estados_actuales) < cantidad_ap:
                    estados_actuales.extend(["Nuevo"] * (cantidad_ap - len(estados_actuales)))
                else:
                    estados_actuales = estados_actuales[:cantidad_ap]
                st.session_state.edificios[i]["estados_aps"] = estados_actuales
        
        with c2:
            tipo = st.selectbox(
                "Crear hojas",
                ["Switches", "AP's", "Ambos"],
                index=["Switches", "AP's", "Ambos"].index(e["tipo"]),
                key=f"tipo_{i}"
            )
            st.session_state.edificios[i]["tipo"] = tipo

            img_tss = st.file_uploader(
                "Imagen TSS (opcional)",
                type=["png", "jpg", "jpeg"],
                key=f"img_tss_{i}"
            )
            if img_tss:
                st.session_state.edificios[i]["imagen_tss"] = img_tss

        # Botón eliminar
        with c3:
            st.write("")
            st.write("")
            if st.button("❌ Eliminar", key=f"del_{i}"):
                st.session_state.edificios.pop(i)
                st.rerun()
        
        # Configuración individual de Switches
        if cantidad_sw > 0 and tipo in ["Switches", "Ambos"]:
            st.write("### 🔌 Configuración de Switches")
            if "estados_switches" not in st.session_state.edificios[i]:
                st.session_state.edificios[i]["estados_switches"] = ["Reemplazo"] * cantidad_sw
            
            cols_sw = st.columns(min(cantidad_sw, 4))
            for sw_idx in range(cantidad_sw):
                with cols_sw[sw_idx % 4]:
                    estado = st.radio(
                        f"Switch {sw_idx + 1}",
                        ["Reemplazo", "Nuevo"],
                        index=0 if st.session_state.edificios[i]["estados_switches"][sw_idx] == "Reemplazo" else 1,
                        key=f"estado_sw_{i}_{sw_idx}",
                        horizontal=True
                    )
                    st.session_state.edificios[i]["estados_switches"][sw_idx] = estado
        
        # Configuración individual de APs
        if cantidad_ap > 0 and tipo in ["AP's", "Ambos"]:
            st.write("### 📡 Configuración de APs")
            if "estados_aps" not in st.session_state.edificios[i]:
                st.session_state.edificios[i]["estados_aps"] = ["Nuevo"] * cantidad_ap
            
            cols_ap = st.columns(min(cantidad_ap, 4))
            for ap_idx in range(cantidad_ap):
                with cols_ap[ap_idx % 4]:
                    estado = st.radio(
                        f"AP {ap_idx + 1}",
                        ["Reemplazo", "Nuevo"],
                        index=0 if st.session_state.edificios[i]["estados_aps"][ap_idx] == "Reemplazo" else 1,
                        key=f"estado_ap_{i}_{ap_idx}",
                        horizontal=True
                    )
                    st.session_state.edificios[i]["estados_aps"][ap_idx] = estado
                    
                    # Imagen para este AP
                    img_ap = st.file_uploader(
                        f"Imagen",
                        type=["png", "jpg", "jpeg"],
                        key=f"img_ap_{i}_{ap_idx}"
                    )
                    if img_ap:
                        if "imagenes_ap" not in st.session_state.edificios[i]:
                            st.session_state.edificios[i]["imagenes_ap"] = [None] * cantidad_ap
                        while len(st.session_state.edificios[i]["imagenes_ap"]) < cantidad_ap:
                            st.session_state.edificios[i]["imagenes_ap"].append(None)
                        st.session_state.edificios[i]["imagenes_ap"][ap_idx] = img_ap


def guardar_seccion_final(ws, fila_inicio_equipos=17):
    """
    Guarda el contenido de la sección final
    """
    seccion_final = []
    fila_busqueda = fila_inicio_equipos
    fila_inicio_seccion = None
    
    for row in range(fila_busqueda, ws.max_row + 1):
        cell_value = ws[f"A{row}"].value
        if cell_value and "INCIDENCIAS" in str(cell_value).upper():
            fila_inicio_seccion = row
            for final_row in range(row, ws.max_row + 1):
                fila_data = {}
                for col in range(1, 9):
                    cell = ws.cell(final_row, col)
                    if not isinstance(cell, MergedCell):
                        fila_data[col] = {
                            'value': cell.value,
                            'font': copy(cell.font) if cell.font else None,
                            'alignment': copy(cell.alignment) if cell.alignment else None,
                            'border': copy(cell.border) if cell.border else None,
                            'fill': copy(cell.fill) if cell.fill else None
                        }
                seccion_final.append(fila_data)
            break
    
    merges_finales = []
    if fila_inicio_seccion:
        for merged_range in list(ws.merged_cells.ranges):
            if merged_range.min_row >= fila_inicio_seccion:
                merges_finales.append({
                    'min_row': merged_range.min_row,
                    'max_row': merged_range.max_row,
                    'min_col': merged_range.min_col,
                    'max_col': merged_range.max_col
                })
    
    return seccion_final, merges_finales, fila_inicio_seccion

def restaurar_seccion_final(ws, seccion_final, merges_finales, fila_destino):
    """
    Restaura la sección final respetando merges complejos
    """
    if not seccion_final:
        return

    # ---- 1. Restaurar SOLO valores (antes de merges) ----
    for idx, fila_data in enumerate(seccion_final):
        row = fila_destino + idx
        for col, cell_data in fila_data.items():
            cell = ws.cell(row=row, column=col)

            # Seguridad extra: nunca escribir sobre MergedCell
            if isinstance(cell, MergedCell):
                continue

            cell.value = cell_data.get("value")

            if cell_data.get("font"):
                cell.font = cell_data["font"]
            if cell_data.get("alignment"):
                cell.alignment = cell_data["alignment"]
            if cell_data.get("border"):
                cell.border = cell_data["border"]
            if cell_data.get("fill"):
                cell.fill = cell_data["fill"]

    # ---- 2. Restaurar merges ----
    if not merges_finales:
        return

    fila_original_inicio = min(m["min_row"] for m in merges_finales)
    desplazamiento = fila_destino - fila_original_inicio

    for merge in merges_finales:
        try:
            r1 = merge["min_row"] + desplazamiento
            r2 = merge["max_row"] + desplazamiento
            c1 = merge["min_col"]
            c2 = merge["max_col"]

            rango = (
                f"{get_column_letter(c1)}{r1}:"
                f"{get_column_letter(c2)}{r2}"
            )

            ws.merge_cells(rango)

        except Exception:
            pass



def limpiar_area_equipos_completa(ws, fila_inicio=17, fila_max=200):

    # 1. Desmergear todo el bloque
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row <= fila_max and merged.max_row >= fila_inicio:
            try:
                ws.unmerge_cells(str(merged))
            except:
                pass

    # 2. Resetear contenido y estilos
    for row in range(fila_inicio, fila_max + 1):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            try:
                cell.value = None
                cell.comment = None
                cell.font = Font()
                cell.alignment = Alignment()
                cell.border = Border()
                cell.fill = PatternFill()
            except:
                pass



def copiar_estructura_switches(ws, num_switches, estados_switches):
    """
    Copia la estructura de la hoja Switches con estados individuales
    """
    fila_inicial = 17
    current_row = fila_inicial
    
    template_labels_sw = [
        "Ubicación (UR)",
        "Tipo",
        "Modelo",
        "Qué equipos se van a migrar (MARCA, NÚMERO DE PUERTOS,UBICACIÓN)",
        "Puertos disponibles para APs",
        "Script(Si se saco el respaldo remotamente)"
    ]
    
    valores_f_switch = ["", "Reemplazo", "Acceso 24 ptos", "", "", "NO"]
    
    for num_sw in range(1, num_switches + 1):
        estado_sw = estados_switches[num_sw - 1] if num_sw - 1 < len(estados_switches) else "Reemplazo"
        
        for idx, label in enumerate(template_labels_sw):
            row = current_row + idx
            
            ws[f"A{row}"].value = f"Switch {num_sw}" if idx == 0 else None
            
            ws[f"B{row}"].value = label
            ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            cell_f = ws[f"F{row}"]
            if idx == 1:
                cell_f.value = estado_sw
            else:
                cell_f.value = valores_f_switch[idx]
            cell_f.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            if num_sw == 1 and idx == 0:
                cell_f.comment = Comment("Reemplazo", "Sistema")
            
            ws[f"G{row}"].value = ""
            
            ws[f"H{row}"].value = "" if idx == 0 else None
        
        ws.merge_cells(f"A{current_row}:A{current_row + 5}")
        ws[f"A{current_row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"A{current_row}"].font = Font(bold=True)
        
        for idx in range(6):
            row = current_row + idx
            ws.merge_cells(f"B{row}:E{row}")
        
        ws.merge_cells(f"H{current_row}:H{current_row + 5}")
        ws[f"H{current_row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        current_row += 6
    
    return current_row


def copiar_estructura_aps(ws, num_aps, estados_aps, imagenes_ap=None):
    """
    Copia la estructura de la hoja AP's con estados individuales
    """
    fila_inicial = 17
    current_row = fila_inicial
    
    template_labels_ap = [
        "Canaleta/Ducteria",
        "Techo falso/Losa",
        "Se reutiliza el cableado",
        "Tipo de cableado",
        "Tipo",
        "Ubicación del AP a instalar",
        "Distancia desde el patch pannel al AP(Aprox)"
    ]
    
    valores_f_ap = ["", "Techo falso", "NO", "", "Nuevo", "", ""]
    
    for num_ap in range(1, num_aps + 1):
        estado_ap = estados_aps[num_ap - 1] if num_ap - 1 < len(estados_aps) else "Nuevo"
        
        for idx, label in enumerate(template_labels_ap):
            row = current_row + idx
            
            ws[f"A{row}"].value = f"AP {num_ap}" if idx == 0 else None
            
            ws[f"B{row}"].value = label
            ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            cell_f = ws[f"F{row}"]
            if idx == 4:
                cell_f.value = estado_ap
            else:
                cell_f.value = valores_f_ap[idx]
            cell_f.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            ws[f"G{row}"].value = ""
            
            ws[f"H{row}"].value = "" if idx == 0 else None
        
        ws.merge_cells(f"A{current_row}:A{current_row + 6}")
        ws[f"A{current_row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"A{current_row}"].font = Font(bold=True)
        
        for idx in range(7):
            row = current_row + idx
            ws.merge_cells(f"B{row}:E{row}")
        
        ws.merge_cells(f"H{current_row}:H{current_row + 6}")
        ws[f"H{current_row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        if imagenes_ap and len(imagenes_ap) >= num_ap and imagenes_ap[num_ap - 1]:
            try:
                img = XLImage(imagenes_ap[num_ap - 1])
                img.width = 180
                img.height = 200
                ws.add_image(img, f"H{current_row}")
            except:
                pass
        
        current_row += 7
    
    return current_row


# --- Generación ---
st.header("4️⃣ Generar Excel por edificio")

if st.button("📥 Generar archivos"):
    if not modelo_file_path or not dependencia or not fecha or not st.session_state.edificios:
        st.error("Faltan datos obligatorios")
    else:

        listado_global = []
        total_sw_r = total_sw_n = total_ap_r = total_ap_n = 0

        for idx, ed in enumerate(st.session_state.edificios, start=1):
            cant_sw = ed.get("cantidad_switches", 1)
            cant_ap = ed.get("cantidad_aps", 1)
            estados_sw = ed.get("estados_switches", [])
            estados_ap = ed.get("estados_aps", [])
            
            # Contar reemplazos y nuevos
            sw_r = sum(1 for i in range(cant_sw) if i < len(estados_sw) and estados_sw[i] == "Reemplazo" and ed["tipo"] in ["Switches", "Ambos"])
            sw_n = sum(1 for i in range(cant_sw) if i < len(estados_sw) and estados_sw[i] == "Nuevo" and ed["tipo"] in ["Switches", "Ambos"])
            
            ap_r = sum(1 for i in range(cant_ap) if i < len(estados_ap) and estados_ap[i] == "Reemplazo" and ed["tipo"] in ["AP's", "Ambos"])
            ap_n = sum(1 for i in range(cant_ap) if i < len(estados_ap) and estados_ap[i] == "Nuevo" and ed["tipo"] in ["AP's", "Ambos"])

            listado_global.append({
                "n": idx,
                "nombre": ed["nombre"],
                "sw_r": sw_r,
                "sw_n": sw_n,
                "ap_r": ap_r,
                "ap_n": ap_n
            })

            total_sw_r += sw_r
            total_sw_n += sw_n
            total_ap_r += ap_r
            total_ap_n += ap_n

        zip_buffer = BytesIO()

        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
            cont = 0

            for e in st.session_state.edificios:
                cont += 1
                wb = load_workbook(modelo_file_path)

                # ---------- TSS ----------
                if "TSS" in wb.sheetnames:
                    ws = wb["TSS"]
                    ws.merge_cells("D10:H10")
                    ws["D10"] = dependencia
                    ws.merge_cells("D11:H11")
                    ws["D11"] = e["nombre"]
                    ws["H6"] = fecha

                    if e.get("imagen_tss"):
                        ws.merge_cells("H17:H32")
                        img = XLImage(e["imagen_tss"])
                        img.width = 180
                        img.height = 360
                        ws.add_image(img, "H17")

                # ---------- SWITCHES ----------
                if "Switches" in wb.sheetnames:
                    if e["tipo"] not in ["Switches", "Ambos"]:
                        wb.remove(wb["Switches"])
                    else:
                        ws = wb["Switches"]
                        ws.merge_cells("D10:H10")
                        ws["D10"] = dependencia
                        ws.merge_cells("D11:H11")
                        ws["D11"] = e["nombre"]
                        ws["H6"] = fecha
                        
                        seccion_final, merges_finales, fila_original = guardar_seccion_final(ws, 17)
                        limpiar_area_equipos_completa(ws, 17)
                        
                        cant_sw = e.get("cantidad_switches", 1)
                        estados_sw = e.get("estados_switches", ["Reemplazo"] * cant_sw)
                        siguiente_fila = copiar_estructura_switches(ws, cant_sw, estados_sw)
                        
                        restaurar_seccion_final(ws, seccion_final, merges_finales, siguiente_fila)

                # ---------- APs ----------
                ws_ap = None
                for name in wb.sheetnames:
                    if name.lower() in ["ap's", "aps", "ap"]:
                        ws_ap = wb[name]
                        break

                if ws_ap:
                    if e["tipo"] not in ["AP's", "Ambos"]:
                        wb.remove(ws_ap)
                    else:
                        ws_ap.merge_cells("D10:H10")
                        ws_ap["D10"] = dependencia
                        ws_ap.merge_cells("D11:H11")
                        ws_ap["D11"] = e["nombre"]
                        ws_ap["H6"] = fecha
                        
                        seccion_final_ap, merges_finales_ap, fila_original_ap = guardar_seccion_final(ws_ap, 17)
                        limpiar_area_equipos_completa(ws_ap, 17)
                        
                        cant_ap = e.get("cantidad_aps", 1)
                        estados_ap = e.get("estados_aps", ["Nuevo"] * cant_ap)
                        imagenes_ap = e.get("imagenes_ap", [])
                        
                        siguiente_fila_ap = copiar_estructura_aps(ws_ap, cant_ap, estados_ap, imagenes_ap)
                        
                        restaurar_seccion_final(ws_ap, seccion_final_ap, merges_finales_ap, siguiente_fila_ap)

                # ---------- LISTADO ----------
                if "Listado" in wb.sheetnames:
                    ws = wb["Listado"]

                    ws["A1"] = "Dependencia:"
                    ws["B1"] = dependencia
                    ws["C1"] = "SW reemplazo"
                    ws["D1"] = "SW nuevo"
                    ws["E1"] = "AP reemplazo"
                    ws["F1"] = "AP nuevo"

                    for col in ["A", "B", "C", "D", "E", "F"]:
                        ws[f"{col}1"].font = Font(bold=True)

                    fila = 2
                    for item in listado_global:
                        ws[f"A{fila}"] = item["n"]
                        ws[f"B{fila}"] = item["nombre"]
                        ws[f"C{fila}"] = item["sw_r"]
                        ws[f"D{fila}"] = item["sw_n"]
                        ws[f"E{fila}"] = item["ap_r"]
                        ws[f"F{fila}"] = item["ap_n"]

                        ws[f"B{fila}"].alignment = Alignment(wrap_text=True)
                        ws[f"A{fila}"].font = Font(bold=True)

                        fila += 1

                    ws[f"A{fila}"] = fila - 1
                    ws[f"B{fila}"] = "TOTAL"
                    ws[f"C{fila}"] = total_sw_r
                    ws[f"D{fila}"] = total_sw_n
                    ws[f"E{fila}"] = total_ap_r
                    ws[f"F{fila}"] = total_ap_n
                    ws[f"B{fila}"].alignment = Alignment(wrap_text=True)

                output = BytesIO()
                wb.save(output)
                output.seek(0)

                nombre = f"{cont}. TSS {dependencia} - {e['nombre']} - MDMQ.xlsx" ######## CAMBIAR EL NOMBRE DEL ARCHIVO #########
                zipf.writestr(nombre, output.read())

        zip_buffer.seek(0)

        st.success("Archivos generados correctamente")

        st.download_button(
            "⬇️ Descargar TODOS los Excel (ZIP)",
            data=zip_buffer,
            file_name=f"TSS {dependencia} - MDMQ.zip",
            mime="application/zip"
        )